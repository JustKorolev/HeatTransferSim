"""Excel part-material lookup loader."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
import re
from typing import Any


@dataclass
class ContactReport:
    component_materials: dict[str, str] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)

    def material_for_component(self, name: str) -> str | None:
        for alias in _component_aliases(name):
            material = self.component_materials.get(alias)
            if material:
                return material
        return None


def load_contact_report(path: str | Path | None) -> ContactReport:
    report = ContactReport()
    if not path:
        return report
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read Excel material lookups.") from exc

    workbook = load_workbook(Path(path), data_only=True, read_only=True)
    try:
        sheet = _find_material_lookup_sheet(workbook)
        _read_material_lookup(sheet, report)
    finally:
        workbook.close()
    _add_component_aliases(report)
    return report


def _header_map(sheet: Any, required: tuple[str, ...]) -> tuple[dict[str, int], int]:
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=50, values_only=True), 1):
        headers = [str(value or "").strip().lower() for value in row]
        if all(name in headers for name in required):
            return {name: index for index, name in enumerate(headers)}, row_number
    raise ValueError(f"Could not find header row in sheet {sheet.title!r}.")


def _pick(row: tuple[Any, ...], headers: dict[str, int], names: tuple[str, ...]) -> Any:
    for name in names:
        index = headers.get(name)
        if index is not None and index < len(row):
            return row[index]
    return None


def _find_material_lookup_sheet(workbook: Any) -> Any:
    if "Materials" in workbook.sheetnames:
        return workbook["Materials"]
    return workbook[workbook.sheetnames[0]]


def _read_material_lookup(sheet: Any, report: ContactReport) -> None:
    headers, header_row = _header_map(sheet, ("part name", "material name"))
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        component = _pick(row, headers, ("part name",))
        if not component:
            continue
        name = str(component).strip()
        material = _pick(row, headers, ("material name",))
        if material:
            report.component_materials[name] = str(material).strip()


def _add_component_aliases(report: ContactReport) -> None:
    candidates: dict[str, set[str]] = {}
    for name, material in report.component_materials.items():
        for alias in _component_aliases(name):
            candidates.setdefault(alias, set()).add(material)
    for alias, materials in candidates.items():
        if alias in report.component_materials or len(materials) != 1:
            continue
        report.component_materials[alias] = next(iter(materials))


def _component_aliases(name: str) -> set[str]:
    aliases = {name}
    current = name
    for _ in range(3):
        stripped = re.sub(r"([_-])\d+$", "", current)
        if stripped == current:
            break
        aliases.add(stripped)
        current = stripped
    return aliases
